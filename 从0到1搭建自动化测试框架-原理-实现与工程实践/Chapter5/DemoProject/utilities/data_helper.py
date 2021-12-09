# -*- coding: utf-8 -*-


import codecs
import json
import os
import yaml
from openpyxl import load_workbook
import pandas as pd


def read_data_from_json_yaml(data_file):
    return_value = []
    data_file_path = os.path.abspath(data_file)
    _is_yaml_file = data_file_path.endswith((".yml", ".yaml"))

    with codecs.open(data_file_path, 'r', 'utf-8') as f:
        # Load the data from YAML or JSON
        if _is_yaml_file:
            data = yaml.safe_load(f)
        else:
            data = json.load(f)
    for i, elem in enumerate(data):
        if isinstance(data, dict):
            key, value = elem, data[elem]
            if isinstance(value, dict):
                case_data = []
                for v in value.values():
                    case_data.append(v)
                return_value.append(tuple(case_data))
            else:
                return_value.append((value,))
    return return_value


def read_data_from_excel(excel_file, sheet_name):
    return_value = []
    if not os.path.exists(excel_file):
        raise ValueError("File not exists")
    wb = load_workbook(excel_file)
    for s in wb.sheetnames:
        if s == sheet_name:
            sheet = wb[sheet_name]
            for row in sheet.rows:
                return_value.append([col.value for col in row])
    return return_value


def read_data_from_pandas(excel_file, sheet_name):
    if not os.path.exists(excel_file):
        raise ValueError("File not exists")
    s = pd.ExcelFile(excel_file)
    df = s.parse(sheet_name)
    return df.values.tolist()
